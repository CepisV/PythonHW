import os
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from openpyxl import load_workbook

def count_rows():
    folder_path = filedialog.askdirectory(title="Выберите папку")
    if not folder_path:
        result_label.config(text="Папка не выбрана")
        return

    total_rows = 0
    for file in os.listdir(folder_path):
        if file.endswith(".xlsx"):
            try:
                workbook = load_workbook(os.path.join(folder_path, file), read_only=True)
                for sheet in workbook.sheetnames:
                    total_rows += workbook[sheet].max_row
            except Exception:
                pass  

    result_label.config(text=f"Всего строк: {total_rows}")

root = tk.Tk()
root.title("Подсчет строк в Excel-файлах")
root.geometry("400x200")
root.resizable(False, False)

style = ttk.Style()
style.configure("TButton", font=("Arial", 12), padding=5)
style.configure("TLabel", font=("Arial", 12))

count_button = ttk.Button(root, text="Выбрать папку и посчитать строки", command=count_rows)
count_button.pack(pady=20)

frame = ttk.Frame(root, padding=10, relief="solid", borderwidth=2)
frame.pack(padx=10, pady=10, fill="both", expand=True)

result_label = ttk.Label(frame, text="Всего строк: 0", anchor="center")
result_label.pack(pady=10)

root.mainloop()
