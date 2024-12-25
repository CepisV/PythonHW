import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
from openpyxl import Workbook, load_workbook
from ttkthemes import ThemedTk


class TaskManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Task Manager")
        self.root.geometry("900x700")
        self.create_database()
        self.setup_styles()
        self.create_gui()

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use("arc")

        self.style.configure("TNotebook", tabposition="nw", background="#ffffff")
        self.style.configure("Treeview", font=("Arial", 12), rowheight=35, background="#f0f0f0", fieldbackground="#f0f0f0")
        self.style.configure("Treeview.Heading", font=("Arial", 13, "bold"), foreground="gray", background="#4CAF50")
        
        self.style.configure("TButton", font=("Arial", 12, "bold"), padding=10, relief="flat", background="#4CAF50", foreground="gray", borderwidth=0)
        self.style.map("TButton", background=[("active", "#45a049")])
        
        self.style.configure("TLabel", font=("Arial", 12, "bold"), foreground="#333333")
        self.style.configure("TEntry", font=("Arial", 12), foreground="#333333")
        self.style.configure("TCombobox", font=("Arial", 12), foreground="#333333")

    def create_database(self):
        self.conn = sqlite3.connect("tasks.db")
        self.cursor = self.conn.cursor()
        self.cursor.execute(""" 
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE
            )
        """)
        self.cursor.execute(""" 
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                description TEXT,
                category_id INTEGER,
                FOREIGN KEY (category_id) REFERENCES categories (id)
            )
        """)
        self.conn.commit()

    def create_gui(self):
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill="both", expand=True)

        self.tasks_tab = ttk.Frame(notebook, style="TFrame")
        self.categories_tab = ttk.Frame(notebook, style="TFrame")

        notebook.add(self.tasks_tab, text="Задачи")
        notebook.add(self.categories_tab, text="Категории")

        self.create_tasks_tab()
        self.create_categories_tab()

    def create_tasks_tab(self):
        self.task_list = ttk.Treeview(
            self.tasks_tab,
            columns=("Title", "Description", "Category"),
            show="headings",
            height=15
        )
        self.task_list.heading("Title", text="Название")
        self.task_list.heading("Description", text="Описание")
        self.task_list.heading("Category", text="Категория")
        self.task_list.pack(side="top", fill="both", padx=10, pady=10, expand=True)

        buttons_frame = tk.Frame(self.tasks_tab, bg="gray")
        buttons_frame.pack(fill="x", padx=10, pady=5)

        ttk.Button(buttons_frame, text="Добавить", command=self.add_task).pack(side="left", padx=5, pady=5)
        ttk.Button(buttons_frame, text="Редактировать", command=self.edit_task).pack(side="left", padx=5, pady=5)
        ttk.Button(buttons_frame, text="Удалить", command=self.delete_task).pack(side="left", padx=5, pady=5)
        ttk.Button(buttons_frame, text="Импорт из Excel", command=self.import_from_excel).pack(side="left", padx=5, pady=5)
        ttk.Button(buttons_frame, text="Сохранить в Excel", command=self.export_to_excel).pack(side="left", padx=5, pady=5)

    def create_categories_tab(self):
        categories_label = ttk.Label(self.categories_tab, text="Управление категориями", font=("Arial", 16, "bold"))
        categories_label.pack(anchor="center", pady=10)

        self.category_list = tk.Listbox(
            self.categories_tab,
            height=15,
            font=("Arial", 12),
            bg="#f9f9f9",
            selectbackground="#4CAF50",
            selectforeground="#ffffff"
        )
        self.category_list.pack(side="top", fill="both", padx=10, pady=10, expand=True)

        buttons_frame = tk.Frame(self.categories_tab, bg="#f0f0f0")
        buttons_frame.pack(fill="x", padx=10, pady=5)

        ttk.Button(buttons_frame, text="Добавить категорию", command=self.add_category).pack(side="left", padx=5, pady=5)
        ttk.Button(buttons_frame, text="Удалить категорию", command=self.delete_category).pack(side="left", padx=5, pady=5)

    def add_task(self):
        add_task_window = tk.Toplevel(self.root)
        add_task_window.title("Добавить задачу")
        add_task_window.geometry("400x300")

        tk.Label(add_task_window, text="Название:", font=("Arial", 12, "bold")).grid(row=0, column=0, padx=10, pady=5)
        title_entry = tk.Entry(add_task_window)
        title_entry.grid(row=0, column=1, padx=10, pady=5)

        tk.Label(add_task_window, text="Описание:", font=("Arial", 12, "bold")).grid(row=1, column=0, padx=10, pady=5)
        description_entry = tk.Entry(add_task_window)
        description_entry.grid(row=1, column=1, padx=10, pady=5)

        tk.Label(add_task_window, text="Категория:", font=("Arial", 12, "bold")).grid(row=2, column=0, padx=10, pady=5)
        category_combobox = ttk.Combobox(add_task_window, values=self.get_categories(), font=("Arial", 12))
        category_combobox.grid(row=2, column=1, padx=10, pady=5)

        def save_task():
            title = title_entry.get()
            description = description_entry.get()
            category = category_combobox.get()

            if not title or not category:
                messagebox.showerror("Ошибка", "Название и категория обязательны для заполнения!")
                return

            self.cursor.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (category,))
            category_id = self.cursor.execute("SELECT id FROM categories WHERE name = ?", (category,)).fetchone()[0]

            self.cursor.execute(
                "INSERT INTO tasks (title, description, category_id) VALUES (?, ?, ?)",
                (title, description, category_id)
            )
            self.conn.commit()
            self.load_tasks()
            add_task_window.destroy()

        tk.Button(add_task_window, text="Сохранить", command=save_task, relief="flat", bg="#4CAF50", fg="white", font=("Arial", 12, "bold")).grid(row=3, columnspan=2, pady=10)

    def edit_task(self):
        selected_task = self.task_list.selection()
        if not selected_task:
            messagebox.showerror("Ошибка", "Выберите задачу для редактирования!")
            return

        task_title = self.task_list.item(selected_task)["values"][0]
        task_description = self.task_list.item(selected_task)["values"][1]
        task_category = self.task_list.item(selected_task)["values"][2]

        edit_task_window = tk.Toplevel(self.root)
        edit_task_window.title("Редактировать задачу")
        edit_task_window.geometry("400x300")

        tk.Label(edit_task_window, text="Название:", font=("Arial", 12, "bold")).grid(row=0, column=0, padx=10, pady=5)
        title_entry = tk.Entry(edit_task_window)
        title_entry.insert(0, task_title)
        title_entry.grid(row=0, column=1, padx=10, pady=5)

        tk.Label(edit_task_window, text="Описание:", font=("Arial", 12, "bold")).grid(row=1, column=0, padx=10, pady=5)
        description_entry = tk.Entry(edit_task_window)
        description_entry.insert(0, task_description)
        description_entry.grid(row=1, column=1, padx=10, pady=5)

        tk.Label(edit_task_window, text="Категория:", font=("Arial", 12, "bold")).grid(row=2, column=0, padx=10, pady=5)
        category_combobox = ttk.Combobox(edit_task_window, values=self.get_categories())
        category_combobox.set(task_category)
        category_combobox.grid(row=2, column=1, padx=10, pady=5)

        def save_edited_task():
            title = title_entry.get()
            description = description_entry.get()
            category = category_combobox.get()

            if not title or not category:
                messagebox.showerror("Ошибка", "Название и категория обязательны для заполнения!")
                return

            category_id = self.cursor.execute("SELECT id FROM categories WHERE name = ?", (category,)).fetchone()[0]

            self.cursor.execute(
                "UPDATE tasks SET title = ?, description = ?, category_id = ? WHERE title = ?",
                (title, description, category_id, task_title)
            )
            self.conn.commit()
            self.load_tasks()
            edit_task_window.destroy()

        tk.Button(edit_task_window, text="Сохранить", command=save_edited_task, relief="flat", bg="#4CAF50", fg="white", font=("Arial", 12, "bold")).grid(row=3, columnspan=2, pady=10)

    def load_tasks(self):
        for row in self.task_list.get_children():
            self.task_list.delete(row)
        
        self.cursor.execute("SELECT tasks.title, tasks.description, categories.name FROM tasks JOIN categories ON tasks.category_id = categories.id")
        tasks = self.cursor.fetchall()
        for task in tasks:
            self.task_list.insert("", "end", values=task)

    def add_category(self):
        category_name = simpledialog.askstring("Новая категория", "Введите название категории:")
        if category_name:
            self.cursor.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (category_name,))
            self.conn.commit()
            self.load_categories()

    def load_categories(self):
        self.category_list.delete(0, tk.END)
        self.cursor.execute("SELECT name FROM categories")
        categories = self.cursor.fetchall()
        for category in categories:
            self.category_list.insert(tk.END, category[0])

    def delete_task(self):
        selected_task = self.task_list.selection()
        if selected_task:
            task_title = self.task_list.item(selected_task)["values"][0]
            confirm = messagebox.askyesno("Подтвердить удаление", f"Вы уверены, что хотите удалить задачу '{task_title}'?")
            if confirm:
                self.cursor.execute("DELETE FROM tasks WHERE title = ?", (task_title,))
                self.conn.commit()
                self.load_tasks()

    def delete_category(self):
        selected_category = self.category_list.curselection()
        if selected_category:
            category_name = self.category_list.get(selected_category)
            confirm = messagebox.askyesno("Подтвердить удаление", f"Вы уверены, что хотите удалить категорию '{category_name}'?")
            if confirm:
                self.cursor.execute("DELETE FROM categories WHERE name = ?", (category_name,))
                self.conn.commit()
                self.load_categories()

    def import_from_excel(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file_path:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                title, description, category = row
                self.cursor.execute("INSERT OR IGNORE INTO categories (name) VALUES (?)", (category,))
                category_id = self.cursor.execute("SELECT id FROM categories WHERE name = ?", (category,)).fetchone()[0]
                self.cursor.execute("INSERT INTO tasks (title, description, category_id) VALUES (?, ?, ?)", (title, description, category_id))
            self.conn.commit()
            self.load_tasks()

    def export_to_excel(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Название", "Описание", "Категория"])
        
        self.cursor.execute("SELECT tasks.title, tasks.description, categories.name FROM tasks JOIN categories ON tasks.category_id = categories.id")
        tasks = self.cursor.fetchall()
        for task in tasks:
            sheet.append(task)

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx;*.xls")])
        if file_path:
            workbook.save(file_path)

    def get_categories(self):
        self.cursor.execute("SELECT name FROM categories")
        categories = [category[0] for category in self.cursor.fetchall()]
        return categories


def main():
    root = ThemedTk()
    app = TaskManagerApp(root)
    app.load_tasks()
    app.load_categories()
    root.mainloop()

if __name__ == "__main__":
    main()
